"""Patch the existing PBD_Losung-Template.xlsx with the Track A improvements.

This script reads ``01-existing-implementation/PBD_Losung-Template.xlsx``,
applies the changes listed in ``02-simple-excel-improvements/README.md``
(A–G), and writes ``PBD_Losung-Patched.xlsx`` next to this script.

What is changed:
  A. New ``GemeindeMapping`` sheet (Nord/Süd assignment of municipalities).
  B. Joint age × sex × region target shares (× education marginal) instead of
     pure product-of-marginals.
  C. Per-cell availability clamp + deficit reallocation to nearest non-empty
     cell.
  D. Hamilton (largest remainders) rounding so the integer counts sum to the
     panel size exactly.
  E. Seeded uniform random draw inside each profile.
  F. Substitutes drawn from the candidate pool minus the chosen members.
  G. New ``Audit`` sheet recording seed, input hash, target vs realised
     marginals, and the drawn IDs.

The original sheets ``KombinationOM`` / ``LosungOM`` / ``KombinationEK`` /
``LosungEK`` are preserved as-is for reference; the new allocation is written
to new sheets ``KombinationOM_v2``, ``LosungOM_v2``, ``KombinationEK_v2``,
``LosungEK_v2`` so the original workbook logic is not disturbed.
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import math
import random
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

ROOT = Path(__file__).resolve().parent
SRC_XLSX = ROOT.parent / "01-existing-implementation" / "PBD_Losung-Template.xlsx"
OUT_XLSX = ROOT / "PBD_Losung-Patched.xlsx"

# Nord / Süd mapping for the 9 municipalities of the German-speaking Community.
# Derived from the totals in the original workbook (the Nord share of 0.6073
# matches Eupen + Kelmis + Lontzen + Raeren).
GEMEINDE_KANTON = {
    "Amel": "Süd",
    "Büllingen": "Süd",
    "Bütgenbach": "Süd",
    "Burg-Reuland": "Süd",
    "St.Vith": "Süd",
    "Eupen": "Nord",
    "Kelmis": "Nord",
    "Lontzen": "Nord",
    "Raeren": "Nord",
}

AGE_BUCKETS = [("16-35", 16, 35), ("36-55", 36, 55), ("56+", 56, 200)]
SEX_VALUES = ["Frau", "Mann"]
REGION_VALUES = ["Nord", "Süd"]
EDU_VALUES = ["GrundschuleLehre", "AbiturMeister", "DualBachelor", "Master"]

PANEL_MEMBERS = 30
PANEL_SUBSTITUTES = 30  # slide deck: 30 substitutes, same modalités

HEADER_FILL = PatternFill("solid", fgColor="FFE2EFDA")
HEADER_FONT = Font(bold=True)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def age_bucket(age: int) -> str:
    for name, lo, hi in AGE_BUCKETS:
        if lo <= age <= hi:
            return name
    raise ValueError(f"age {age} outside considered range")


def hamilton(fractional: dict, total: int) -> dict:
    """Largest-remainders (Hamilton) rounding.

    Given a mapping ``key -> fractional_count``, return integer counts summing
    to ``total``. Ties on the remainder are broken by a stable key order.
    """
    floors = {k: int(math.floor(v)) for k, v in fractional.items()}
    assigned = sum(floors.values())
    remaining = total - assigned
    if remaining == 0:
        return floors
    if remaining < 0:
        # Defensive: floors over-shot total (only possible with non-standard
        # fractional inputs). Shave from the *smallest* remainders first —
        # those cells "deserved" the floor the least, so taking the +1 away
        # is the least disruptive choice.
        order = sorted(
            fractional, key=lambda k: (fractional[k] - floors[k], k)
        )
        for k in order[: -remaining]:
            floors[k] -= 1
        return floors
    order = sorted(fractional, key=lambda k: (-(fractional[k] - floors[k]), k))
    for k in order[:remaining]:
        floors[k] += 1
    return floors


def hash_candidates(rows: list[tuple]) -> str:
    h = hashlib.sha256()
    for row in rows:
        h.update("|".join("" if v is None else str(v) for v in row).encode("utf-8"))
        h.update(b"\n")
    return h.hexdigest()


def neighbour_order(profile: tuple[str, str, str, str]) -> list:
    """Return a list of partial-match keys ordered from most to least specific.

    Used when a cell is empty / oversubscribed and the deficit/surplus has to
    be moved to the closest non-empty cell. The order is: same (sex, age,
    region), then (sex, age), then (sex, region), then (age, region), then
    (sex,), (age,), (region,), then any.
    """
    sex, age, region, _ = profile
    return [
        ("sex_age_region", (sex, age, region)),
        ("sex_age", (sex, age)),
        ("sex_region", (sex, region)),
        ("age_region", (age, region)),
        ("sex", (sex,)),
        ("age", (age,)),
        ("region", (region,)),
        ("any", ()),
    ]


# ---------------------------------------------------------------------------
# Read inputs
# ---------------------------------------------------------------------------


def read_candidates(wb) -> list[dict]:
    ws = wb["VerfügbareTeilnehmer"]
    out = []
    for r in range(3, ws.max_row + 1):
        cid = ws.cell(r, 1).value
        if not cid:
            continue
        out.append(
            {
                "ID": cid,
                "Geschlecht": ws.cell(r, 2).value,
                "Alterskategorie": ws.cell(r, 3).value,
                "Kanton": ws.cell(r, 4).value,
                "Ausbildung": ws.cell(r, 5).value,
                "Profil": ws.cell(r, 6).value,
            }
        )
    return out


def read_population_joint(wb) -> dict:
    """Return joint counts P[(sex, age_bucket, region)] from Bevölkerungsstruktur.

    The sheet is laid out as: rows = individual ages 16..100, columns grouped
    by municipality, then by nationality (Belgier / Nicht Belgier), then by
    sex (Männer / Frauen). The four nationality-sex sub-columns per
    municipality are at offsets 0..3 starting at column 2 for Amel.
    """
    ws = wb["Bevölkerungsstruktur"]
    # Column layout (1-indexed): col 1 is age; municipalities start at col 2
    # and each takes 4 columns (Belgier Männer, Belgier Frauen, Nicht Männer,
    # Nicht Frauen). Order matches GEMEINDE_KANTON keys in the original sheet.
    gemeinde_order = [
        "Amel",
        "Büllingen",
        "Bütgenbach",
        "Burg-Reuland",
        "St.Vith",
        "Eupen",
        "Kelmis",
        "Lontzen",
        "Raeren",
    ]
    joint = defaultdict(int)
    education_marginal = {}
    # Education marginal lives in columns AN/AO (40/41) at rows 2..7
    # Row 2 = "Ohne Diplom", we map the 6 official categories to the workbook's
    # 4 attribute values using the same grouping the original workbook uses:
    #   GrundschuleLehre = Ohne Diplom + Primarschule + Sekundarschule Unterstufe
    #   AbiturMeister    = Sekundarschule Oberstufe
    #   DualBachelor     = Hochschule kurzer Studiengang
    #   Master           = Hochschule langer Studiengang + Universität
    edu_raw = {}
    for r in range(2, 10):
        label = ws.cell(r, 40).value
        val = ws.cell(r, 41).value
        if label and isinstance(val, (int, float)):
            edu_raw[str(label).strip()] = val
    education_marginal["GrundschuleLehre"] = (
        edu_raw.get("Ohne Diplom", 0)
        + edu_raw.get("Primarschule", 0)
        + edu_raw.get("Sekundarschule, Unterstufe", 0)
    )
    education_marginal["AbiturMeister"] = edu_raw.get("Sekundarschule, Oberstufe", 0)
    education_marginal["DualBachelor"] = edu_raw.get("Hochschule, kurzer Studiengang", 0)
    education_marginal["Master"] = edu_raw.get("Hochschule, langer Studiengang", 0) + edu_raw.get(
        "Universität", 0
    )

    for r in range(5, 90):
        age = ws.cell(r, 1).value
        if not isinstance(age, (int, float)):
            continue
        bucket = age_bucket(int(age))
        for gi, gemeinde in enumerate(gemeinde_order):
            base = 2 + gi * 4
            b_m = ws.cell(r, base).value or 0
            b_f = ws.cell(r, base + 1).value or 0
            n_m = ws.cell(r, base + 2).value or 0
            n_f = ws.cell(r, base + 3).value or 0
            region = GEMEINDE_KANTON[gemeinde]
            joint[("Mann", bucket, region)] += b_m + n_m
            joint[("Frau", bucket, region)] += b_f + n_f
    return joint, education_marginal


# ---------------------------------------------------------------------------
# Allocation
# ---------------------------------------------------------------------------


def compute_targets(joint: dict, edu_marginal: dict, panel_size: int) -> dict:
    """Return integer target counts per (sex, age, region, edu) profile.

    Joint share comes from age × sex × region cross-tab; education is applied
    as a marginal correction (assumed independent of the first three because
    only marginal data is available — better than full independence on all
    four).
    """
    joint_total = sum(joint.values())
    edu_total = sum(edu_marginal.values())
    fractional = {}
    for sex in SEX_VALUES:
        for _, bucket_name in [(b[0], b[0]) for b in AGE_BUCKETS]:
            for region in REGION_VALUES:
                p_joint = joint[(sex, bucket_name, region)] / joint_total
                for edu in EDU_VALUES:
                    p_edu = edu_marginal[edu] / edu_total
                    fractional[(sex, bucket_name, region, edu)] = panel_size * p_joint * p_edu
    return hamilton(fractional, panel_size), fractional


def clamp_and_reallocate(targets: dict, available: dict) -> tuple[dict, list]:
    """Clamp each target by availability; redistribute deficit to neighbours.

    Returns ``(adjusted_targets, log)`` where ``log`` contains one entry per
    moved seat, in the form ``(from_profile, to_profile, reason)``.
    """
    adjusted = {k: min(v, available.get(k, 0)) for k, v in targets.items()}
    deficit_pool = sum(targets.values()) - sum(adjusted.values())
    log: list = []
    if deficit_pool == 0:
        return adjusted, log

    # Available remaining capacity per profile (after the clamp)
    remaining = {k: available.get(k, 0) - adjusted[k] for k in adjusted}

    # Distribute each deficit seat to the closest profile with spare capacity.
    # For each empty/clamped cell that produced deficit, try its neighbours in
    # decreasing specificity. Process cells with the largest deficit first to
    # keep behaviour deterministic.
    deficits = []
    for k, t in targets.items():
        d = t - adjusted[k]
        if d > 0:
            deficits.append((k, d))
    deficits.sort(key=lambda x: (-x[1], x[0]))

    for profile, d in deficits:
        for _ in range(d):
            placed = False
            for level, key in neighbour_order(profile):
                # find a profile that matches `key` at this specificity level
                # and has remaining capacity. Prefer the one with the highest
                # remaining capacity to spread the load.
                candidates = []
                for other in remaining:
                    if remaining[other] <= 0 or other == profile:
                        continue
                    other_key = {
                        "sex_age_region": (other[0], other[1], other[2]),
                        "sex_age": (other[0], other[1]),
                        "sex_region": (other[0], other[2]),
                        "age_region": (other[1], other[2]),
                        "sex": (other[0],),
                        "age": (other[1],),
                        "region": (other[2],),
                        "any": (),
                    }[level]
                    if other_key == key:
                        candidates.append(other)
                if not candidates:
                    continue
                candidates.sort(key=lambda o: (-remaining[o], o))
                target = candidates[0]
                adjusted[target] += 1
                remaining[target] -= 1
                log.append((profile, target, level))
                placed = True
                break
            if not placed:
                log.append((profile, None, "unplaced"))
    return adjusted, log


def draw_panel(
    candidates: list[dict], targets: dict, seed: int
) -> tuple[list[dict], list[dict]]:
    """Return (drawn, remaining) for the given per-profile targets.

    Within each profile, ``target`` candidates are drawn uniformly at random
    using ``random.Random(seed)``.
    """
    by_profile: dict = defaultdict(list)
    for c in candidates:
        key = (c["Geschlecht"], c["Alterskategorie"], c["Kanton"], c["Ausbildung"])
        by_profile[key].append(c)
    # Sort each per-profile pool by ID so the draw is independent of the
    # row order in the source workbook (and therefore reproducible after
    # any local Excel re-sort).
    for pool in by_profile.values():
        pool.sort(key=lambda c: str(c["ID"]))

    rng = random.Random(seed)
    drawn: list[dict] = []
    drawn_ids: set = set()
    for profile, t in sorted(targets.items()):
        pool = by_profile.get(profile, [])
        if t <= 0 or not pool:
            continue
        chosen = rng.sample(pool, min(t, len(pool)))
        drawn.extend(chosen)
        drawn_ids.update(c["ID"] for c in chosen)

    remaining = [c for c in candidates if c["ID"] not in drawn_ids]
    return drawn, remaining


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------


def write_header(ws, row: int, values: list[str]) -> None:
    for c, v in enumerate(values, start=1):
        cell = ws.cell(row, c, v)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL


def write_gemeinde_mapping(wb) -> None:
    name = "GemeindeMapping"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    write_header(ws, 1, ["Gemeinde", "Kanton"])
    for i, (g, k) in enumerate(GEMEINDE_KANTON.items(), start=2):
        ws.cell(i, 1, g)
        ws.cell(i, 2, k)


def write_kombination(wb, sheet_name: str, fractional: dict, targets: dict,
                      adjusted: dict, available: dict) -> None:
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    write_header(
        ws,
        1,
        [
            "Geschlecht",
            "Alterskategorie",
            "Kanton",
            "Ausbildung",
            "Verfügbar",
            "Ziel (fraktional)",
            "Ziel (Hamilton)",
            "Ziel (geklemmt)",
            "Differenz",
        ],
    )
    row = 2
    for key in sorted(targets, key=lambda k: (-targets[k], k)):
        sex, age, region, edu = key
        ws.cell(row, 1, sex)
        ws.cell(row, 2, age)
        ws.cell(row, 3, region)
        ws.cell(row, 4, edu)
        ws.cell(row, 5, available.get(key, 0))
        ws.cell(row, 6, round(fractional[key], 4))
        ws.cell(row, 7, targets[key])
        ws.cell(row, 8, adjusted[key])
        ws.cell(row, 9, adjusted[key] - targets[key])
        row += 1
    ws.cell(row + 1, 4, "Summe")
    ws.cell(row + 1, 7, sum(targets.values()))
    ws.cell(row + 1, 8, sum(adjusted.values()))


def write_losung(wb, sheet_name: str, drawn: list[dict]) -> None:
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    write_header(
        ws,
        1,
        ["ID", "Geschlecht", "Alterskategorie", "Kanton", "Ausbildung", "Profil"],
    )
    for i, c in enumerate(drawn, start=2):
        ws.cell(i, 1, c["ID"])
        ws.cell(i, 2, c["Geschlecht"])
        ws.cell(i, 3, c["Alterskategorie"])
        ws.cell(i, 4, c["Kanton"])
        ws.cell(i, 5, c["Ausbildung"])
        ws.cell(i, 6, c["Profil"])


def write_audit(
    wb,
    seed: int,
    candidate_hash: str,
    targets_om: dict,
    adjusted_om: dict,
    drawn_om: list[dict],
    targets_ek: dict,
    adjusted_ek: dict,
    drawn_ek: list[dict],
    reallocation_log_om: list,
    reallocation_log_ek: list,
) -> None:
    name = "Audit"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws.cell(1, 1, "Zeitpunkt").font = HEADER_FONT
    ws.cell(1, 2, dt.datetime.now().isoformat(timespec="seconds"))
    ws.cell(2, 1, "Seed").font = HEADER_FONT
    ws.cell(2, 2, seed)
    ws.cell(3, 1, "Kandidaten-Hash (SHA-256)").font = HEADER_FONT
    ws.cell(3, 2, candidate_hash)
    ws.cell(4, 1, "Generator").font = HEADER_FONT
    ws.cell(4, 2, "patch_workbook.py (Track A)")

    def marginals(items: list[dict]) -> dict:
        m: dict = defaultdict(int)
        for c in items:
            m[("Geschlecht", c["Geschlecht"])] += 1
            m[("Alterskategorie", c["Alterskategorie"])] += 1
            m[("Kanton", c["Kanton"])] += 1
            m[("Ausbildung", c["Ausbildung"])] += 1
        return m

    row = 6
    for title, drawn, adjusted in [
        ("Ordentliche Mitglieder", drawn_om, adjusted_om),
        ("Ersatzkandidaten", drawn_ek, adjusted_ek),
    ]:
        ws.cell(row, 1, title).font = HEADER_FONT
        row += 1
        write_header(row + 0 - 1, 1, []) if False else None  # noqa
        write_header(ws, row, ["Attribut", "Wert", "Soll (Summe)", "Ist (gezogen)"])
        row += 1
        target_marginals: dict = defaultdict(int)
        for key, v in adjusted.items():
            sex, age, region, edu = key
            target_marginals[("Geschlecht", sex)] += v
            target_marginals[("Alterskategorie", age)] += v
            target_marginals[("Kanton", region)] += v
            target_marginals[("Ausbildung", edu)] += v
        m = marginals(drawn)
        keys = sorted(set(list(target_marginals) + list(m)))
        for attr, val in keys:
            ws.cell(row, 1, attr)
            ws.cell(row, 2, val)
            ws.cell(row, 3, target_marginals.get((attr, val), 0))
            ws.cell(row, 4, m.get((attr, val), 0))
            row += 1
        row += 1

    ws.cell(row, 1, "Umverteilte Sitze (Defizit-Reallokation)").font = HEADER_FONT
    row += 1
    write_header(ws, row, ["Liste", "Von Profil", "Auf Profil", "Ebene"])
    row += 1
    for label, log in [("OM", reallocation_log_om), ("EK", reallocation_log_ek)]:
        for src, dst, level in log:
            ws.cell(row, 1, label)
            ws.cell(row, 2, "/".join(src) if src else "")
            ws.cell(row, 3, "/".join(dst) if dst else "(unplaced)")
            ws.cell(row, 4, level)
            row += 1


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--seed", type=int, default=20260527,
                        help="Random seed for the draw (default: 20260527).")
    parser.add_argument("--members", type=int, default=PANEL_MEMBERS)
    parser.add_argument("--substitutes", type=int, default=PANEL_SUBSTITUTES)
    parser.add_argument("--source", type=Path, default=SRC_XLSX)
    parser.add_argument("--output", type=Path, default=OUT_XLSX)
    args = parser.parse_args()

    print(f"Reading {args.source} …")
    wb = openpyxl.load_workbook(args.source, data_only=True)
    candidates = read_candidates(wb)
    print(f"  {len(candidates)} candidates loaded")
    joint, edu_marginal = read_population_joint(wb)
    print(f"  population joint cells (sex × age × region): {len(joint)}")

    candidate_hash = hash_candidates(
        [(c["ID"], c["Geschlecht"], c["Alterskategorie"], c["Kanton"], c["Ausbildung"])
         for c in candidates]
    )

    # Per-profile availability
    available: dict = defaultdict(int)
    for c in candidates:
        available[(c["Geschlecht"], c["Alterskategorie"], c["Kanton"], c["Ausbildung"])] += 1

    # --- Members (OM) ----------------------------------------------------
    targets_om_int, fractional_om = compute_targets(joint, edu_marginal, args.members)
    adjusted_om, log_om = clamp_and_reallocate(targets_om_int, available)
    drawn_om, remaining = draw_panel(candidates, adjusted_om, seed=args.seed)
    print(f"  drew {len(drawn_om)} members")

    # --- Substitutes (EK) from the remainder -----------------------------
    available_ek: dict = defaultdict(int)
    for c in remaining:
        available_ek[(c["Geschlecht"], c["Alterskategorie"], c["Kanton"], c["Ausbildung"])] += 1
    targets_ek_int, fractional_ek = compute_targets(joint, edu_marginal, args.substitutes)
    adjusted_ek, log_ek = clamp_and_reallocate(targets_ek_int, available_ek)
    drawn_ek, _ = draw_panel(remaining, adjusted_ek, seed=args.seed + 1)
    print(f"  drew {len(drawn_ek)} substitutes")

    # --- Write everything ------------------------------------------------
    write_gemeinde_mapping(wb)
    write_kombination(wb, "KombinationOM_v2", fractional_om, targets_om_int,
                      adjusted_om, available)
    write_losung(wb, "LosungOM_v2", drawn_om)
    write_kombination(wb, "KombinationEK_v2", fractional_ek, targets_ek_int,
                      adjusted_ek, available_ek)
    write_losung(wb, "LosungEK_v2", drawn_ek)
    write_audit(wb, args.seed, candidate_hash,
                targets_om_int, adjusted_om, drawn_om,
                targets_ek_int, adjusted_ek, drawn_ek,
                log_om, log_ek)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"Wrote {args.output}")


if __name__ == "__main__":
    main()
