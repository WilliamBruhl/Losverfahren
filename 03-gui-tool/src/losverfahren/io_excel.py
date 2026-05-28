# SPDX-License-Identifier: AGPL-3.0-or-later
"""Read the candidate list and population reference from the existing workbook.

The workbook shape is the one of ``PBD_Losung-Template.xlsx``:

* ``VerfügbareTeilnehmer`` — candidate rows starting at row 3, columns
  A-F: ID, Geschlecht, Alterskategorie, Kanton, Ausbildung, Profil.
* ``Bevölkerungsstruktur`` — age × sex × municipality × nationality
  cross-tab; education marginal in columns AN/AO at rows 2..7.
"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

import openpyxl


GEMEINDE_KANTON: dict[str, str] = {
    "Amel": "Süd",
    "Büllingen": "Süd",
    "Bütgenbach": "Süd",
    "Burg-Reuland": "Süd",
    "St.Vith": "Süd",
    "Eupen": "Nord",
    "Kelmis": "Nord",
    "Lontzen": "Nord",
    "Raeren": "Nord",
}

AGE_BUCKETS: list[tuple[str, int, int]] = [
    ("16-35", 16, 35),
    ("36-55", 36, 55),
    ("56+", 56, 200),
]

# Mapping used by the original workbook to collapse 6 official education
# categories onto its 4 attribute values.
EDU_GROUPING: dict[str, list[str]] = {
    "GrundschuleLehre": ["Ohne Diplom", "Primarschule", "Sekundarschule, Unterstufe"],
    "AbiturMeister": ["Sekundarschule, Oberstufe"],
    "DualBachelor": ["Hochschule, kurzer Studiengang"],
    "Master": ["Hochschule, langer Studiengang", "Universität"],
}

# Legacy column order from the original PBD workbook. Other parts of the code
# no longer hard‑depend on this list — the active feature set is now derived
# at runtime from the intersection of candidates and population. The constant
# is kept only as the default for the legacy Excel reader / writer.
LEGACY_FEATURES = ["Geschlecht", "Alterskategorie", "Kanton", "Ausbildung"]
FEATURES = LEGACY_FEATURES  # backwards‑compatible alias


@dataclass
class Candidate:
    """One candidate with an arbitrary set of stratification attributes.

    ``attrs`` is the source of truth (a flat ``feature → value`` dict). Any
    column name in ``candidates.csv`` other than ``ID`` is preserved here
    verbatim, so adding a new field such as ``Beruf`` requires no code
    change. ``Profil``, if present in legacy inputs, is treated as a regular
    attribute but ignored by the solver unless it also appears in the
    population file.
    """

    ID: str
    attrs: dict[str, str] = field(default_factory=dict)

    # ---- backwards‑compatibility shims ------------------------------
    # Older call sites read ``c.Geschlecht`` etc. as plain attributes. We
    # surface every key in ``attrs`` the same way, so existing code keeps
    # working without touching every reference.
    def __getattr__(self, name: str) -> str | None:  # pragma: no cover - shim
        if name.startswith("_") or name in {"ID", "attrs"}:
            raise AttributeError(name)
        attrs = self.__dict__.get("attrs") or {}
        if name in attrs:
            return attrs[name]
        # ``Profil`` was previously a typed field defaulting to ``None``.
        if name == "Profil":
            return None
        raise AttributeError(name)


def _age_bucket(age: int) -> str:
    for name, lo, hi in AGE_BUCKETS:
        if lo <= age <= hi:
            return name
    raise ValueError(f"age {age} out of range")


def read_candidates(path: str | Path) -> list[Candidate]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["VerfügbareTeilnehmer"]
    out: list[Candidate] = []
    for r in range(3, ws.max_row + 1):
        cid = ws.cell(r, 1).value
        if not cid:
            continue
        attrs = {
            "Geschlecht": str(ws.cell(r, 2).value),
            "Alterskategorie": str(ws.cell(r, 3).value),
            "Kanton": str(ws.cell(r, 4).value),
            "Ausbildung": str(ws.cell(r, 5).value),
        }
        profil = ws.cell(r, 6).value
        if profil:
            attrs["Profil"] = str(profil)
        out.append(Candidate(ID=str(cid), attrs=attrs))
    return out


def read_population_marginals(path: str | Path) -> dict[str, dict[str, float]]:
    """Return marginal population *shares* per feature value.

    Joint (age × sex × region) shares are computed from the real cross-tab in
    ``Bevölkerungsstruktur`` and then collapsed to the three marginals; the
    education share comes from the marginal in columns AN/AO.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Bevölkerungsstruktur"]

    gemeinde_order = list(GEMEINDE_KANTON.keys())
    joint: dict[tuple[str, str, str], int] = defaultdict(int)
    for r in range(5, ws.max_row + 1):
        age = ws.cell(r, 1).value
        if not isinstance(age, (int, float)):
            continue
        bucket = _age_bucket(int(age))
        for gi, gemeinde in enumerate(gemeinde_order):
            base = 2 + gi * 4
            b_m = ws.cell(r, base).value or 0
            b_f = ws.cell(r, base + 1).value or 0
            n_m = ws.cell(r, base + 2).value or 0
            n_f = ws.cell(r, base + 3).value or 0
            region = GEMEINDE_KANTON[gemeinde]
            joint[("Mann", bucket, region)] += b_m + n_m
            joint[("Frau", bucket, region)] += b_f + n_f

    total = sum(joint.values()) or 1
    sex = defaultdict(int)
    age = defaultdict(int)
    region = defaultdict(int)
    for (s, a, r), v in joint.items():
        sex[s] += v
        age[a] += v
        region[r] += v

    edu_raw: dict[str, float] = {}
    for r in range(2, 10):
        label = ws.cell(r, 40).value
        val = ws.cell(r, 41).value
        if label and isinstance(val, (int, float)):
            edu_raw[str(label).strip()] = float(val)
    edu = {k: sum(edu_raw.get(s, 0.0) for s in srcs) for k, srcs in EDU_GROUPING.items()}
    edu_total = sum(edu.values()) or 1

    return {
        "Geschlecht": {k: v / total for k, v in sex.items()},
        "Alterskategorie": {k: v / total for k, v in age.items()},
        "Kanton": {k: v / total for k, v in region.items()},
        "Ausbildung": {k: v / edu_total for k, v in edu.items()},
    }


def write_result_workbook(
    src_path: str | Path | None,
    out_path: str | Path,
    members: list[Candidate],
    substitutes: list[Candidate],
    probabilities: dict[str, float],
    audit_rows: list[tuple[str, str]],
) -> None:
    """Save a workbook with members, substitutes, probabilities, audit.

    If ``src_path`` is given (legacy Excel workflow), the original sheets are
    preserved and the new ``LosungLP_*`` sheets are appended. If ``src_path``
    is ``None`` (CSV workflow), a fresh workbook is created.
    """

    if src_path is not None:
        wb = openpyxl.load_workbook(src_path, data_only=True)
    else:
        wb = openpyxl.Workbook()
        # remove default sheet so we only have the LosungLP_* ones
        default = wb.active
        wb.remove(default)

    for name in ("LosungLP_Mitglieder", "LosungLP_Ersatz", "LosungLP_Audit",
                 "LosungLP_Probabilities"):
        if name in wb.sheetnames:
            del wb[name]

    def _write_panel(name: str, items: Iterable[Candidate]) -> None:
        items = list(items)
        ws = wb.create_sheet(name)
        # Discover columns dynamically: ID + union of all attribute keys,
        # preserving order of first appearance so the legacy column order is
        # kept when the inputs follow it.
        cols: list[str] = []
        seen: set[str] = set()
        for c in items:
            for k in c.attrs.keys():
                if k not in seen:
                    cols.append(k)
                    seen.add(k)
        ws.append(["ID", *cols])
        for c in items:
            ws.append([c.ID, *(c.attrs.get(k, "") for k in cols)])

    _write_panel("LosungLP_Mitglieder", members)
    _write_panel("LosungLP_Ersatz", substitutes)

    ws_p = wb.create_sheet("LosungLP_Probabilities")
    ws_p.append(["ID", "p_member", "p_substitute"])
    # Probabilities dict is flat — we store the member probability; the
    # substitute draw uses a separate LP and is logged via the panel sheet.
    for cid, p in sorted(probabilities.items()):
        ws_p.append([cid, round(p, 6), ""])

    ws_a = wb.create_sheet("LosungLP_Audit")
    for k, v in audit_rows:
        ws_a.append([k, v])

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
